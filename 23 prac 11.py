from openpyxl import load_workbook

wb = load_workbook("exc.xlsx", read_only=True)
ws = wb.active

# Read specific range
for row in ws.iter_rows(min_row=1, max_row=10, max_col=5, values_only=True):
    print(row)
