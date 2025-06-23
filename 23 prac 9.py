from openpyxl import load_workbook

# Load the workbook
wb = load_workbook(r'C:\Users\MANOJ\OneDrive\Documents\exc.xlsx', read_only=True)  # Use read_only mode for large files
ws = wb.active

# Read rows one by one
for row in ws.iter_rows(min_row=1, max_row=10, max_col=2, values_only=True):
    print(row)  # Print the first 10 rows
