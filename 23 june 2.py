from openpyxl import Workbook,load_workbook
wb = Workbook()
sheet = wb.active
sheet.title = "Sheet1"
sheet["A1"] = "Name"
sheet["A2"] = "Age"
sheet.append(["Job","23"])
sheet.append(["Alice","25"])
wb.save(r"C:\Users\MANOJ\OneDrive\Documents\excelpy.xlsx")

wb = load_workbook(r"C:\Users\MANOJ\OneDrive\Documents\excelpy.xlsx")
sheet = wb.active
for row in sheet.iter_rows(values_only=True):
    sheet.iter_rows(values_only=True)
    print(row)
